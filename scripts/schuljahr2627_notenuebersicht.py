#!/usr/bin/env python3
"""Baut aus den School-SH-Noten-Exporten (export_schueler_*.xls) je Klasse eine kompakte
Notenuebersicht: Nachname, Vorname, gewichteter Notendurchschnitt (Hauptfaecher 3-fach), und bis zu
3 auffaellige Einzelnoten (1er oder 5er/6er). Eine Ausgabedatei pro Klasse, im jeweiligen Klassenordner.

Hauptfaecher (3-fach gewichtet): Deutsch, Mathematik, Englisch - Standardannahme fuer Sek I,
kann bei Bedarf angepasst werden. Alle anderen Fachnoten zaehlen einfach.
Genutzt wird jeweils die letzte VERFUEGBARE Note (2. Halbjahr 25/26 - 26/27 ist noch leer).
"""
import glob
import xlrd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from pathlib import Path

BASE = "/mnt/d/OneDrive/Dokumente/Office/7. Klassen"
HAUPTFAECHER = {"Deutsch", "Mathematik", "Englisch"}
KURZNAMEN = {
    "Deutsch": "Deutsch", "Mathematik": "Mathe", "Englisch": "Englisch",
    "Französisch": "Franz", "Biologie": "Bio", "Physik": "Physik",
    "Geschichte": "Geschichte", "Geographie": "Geo", "Philosophie": "Philo",
    "Religion": "Reli", "Kunst": "Kunst", "Musik": "Musik", "Sport": "Sport",
}
SKIP_SPALTEN = {"Name", "Versäumnis", "Jahrgangsstufe", "Versetzung",
                "Empfehlung / Prognose", "Überfachliche Kompetenzen", "Zeugnisbemerkungen"}


def clean_subject(raw):
    """'Deutsch (Ar)' -> 'Deutsch'"""
    return raw.split(" (")[0].strip()


def parse_klasse(xls_path):
    book = xlrd.open_workbook(xls_path)
    sheet = book.sheet_by_index(0)

    hdr_row = next(r for r in range(sheet.nrows) if sheet.cell_value(r, 0) == "Name")

    # Fach-Spalten ermitteln: jede nicht-leere Kopfzeile ab Spalte 5 markiert den Start eines
    # 3er-Blocks (25/26 2.HJ | 26/27 1.HJ | 26/27 2.HJ), ausser den beiden Text-Spalten am Ende.
    subjects = []  # (clean_name, col_index)
    for c in range(5, sheet.ncols):
        raw = sheet.cell_value(hdr_row, c)
        if not raw or raw in SKIP_SPALTEN:
            continue
        subjects.append((clean_subject(raw), c))

    students = []
    r = hdr_row + 3
    while r < sheet.nrows and sheet.cell_value(r, 0):
        name_raw = sheet.cell_value(r, 0)
        if ", " in name_raw:
            nachname, vorname = name_raw.split(", ", 1)
        else:
            nachname, vorname = name_raw, ""

        grades = []  # (subject, grade_int)
        for subj, col in subjects:
            val = sheet.cell_value(r, col)
            try:
                g = int(val)
                if 1 <= g <= 6:
                    grades.append((subj, g))
            except (ValueError, TypeError):
                continue  # '/' oder leer = Fach nicht belegt

        weighted_sum = sum(g * (3 if subj in HAUPTFAECHER else 1) for subj, g in grades)
        weight_total = sum(3 if subj in HAUPTFAECHER else 1 for subj, _ in grades)
        avg = round(weighted_sum / weight_total, 2) if weight_total else None

        auffaellig = [(subj, g) for subj, g in grades if g == 1 or g >= 5]
        # Reihenfolge: erst die besten (1er), dann die schlechtesten (5/6, schlechteste zuerst)
        auffaellig.sort(key=lambda x: (x[1] != 1, -x[1] if x[1] >= 5 else 0))
        auffaellig_str = ", ".join(f"{KURZNAMEN.get(subj, subj)} {g}" for subj, g in auffaellig[:3])

        students.append({
            "nachname": nachname, "vorname": vorname,
            "durchschnitt": avg, "auffaellig": auffaellig_str,
        })
        r += 1

    return students


def write_output(klasse, students, out_path):
    wb = Workbook()
    ws = wb.active
    ws.title = f"Noten {klasse}"
    headers = ["Nachname", "Vorname", "Durchschnitt", "Auffällige Noten"]
    ws.append(headers)
    for c in range(1, 5):
        ws.cell(row=1, column=c).font = Font(bold=True)
    for s in students:
        ws.append([s["nachname"], s["vorname"], s["durchschnitt"], s["auffaellig"]])
    for row in ws.iter_rows(min_row=2, min_col=3, max_col=3):
        for cell in row:
            cell.number_format = "0.00"
    widths = [18, 16, 13, 30]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w
    ws.freeze_panes = "A2"
    wb.save(out_path)


if __name__ == "__main__":
    for klasse in ["7a", "7b", "7c", "7d"]:
        matches = glob.glob(f"{BASE}/{klasse}/export_schueler_*.xls")
        if not matches:
            print(f"⚠️ {klasse}: keine export_schueler_*.xls gefunden, übersprungen")
            continue
        students = parse_klasse(matches[0])
        out_path = Path(BASE) / klasse / f"Notenuebersicht_{klasse}.xlsx"
        write_output(klasse, students, out_path)
        n_avg = sum(1 for s in students if s["durchschnitt"] is not None)
        print(f"✅ {klasse}: {len(students)} Schüler, {n_avg} mit Durchschnitt → {out_path}")
